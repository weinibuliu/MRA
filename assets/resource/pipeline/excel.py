from maa.toolkit import Toolkit
from maa.context import Context
from maa.custom_action import CustomAction
from maa.custom_recognition import CustomRecognition
import openpyxl  # 用于操作Excel文件

def main():
    # 注册自定义动作
    Toolkit.pi_register_custom_action("OCRToExcel", OCRToExcelAction())
    Toolkit.pi_register_custom_action("OCRCity", OCRCityAction())

    # 启动 MaaPiCli
    Toolkit.pi_run_cli(r"D:\MRA\assets", r"D:\MRA\cache", False)

class OCRToExcelAction(CustomAction):
    def run(
        self, context: Context, argv: CustomAction.RunArg
    ) -> bool:

        print(f"正在执行 OCRToExcelAction.run，context: {context}, argv: {argv}")
        # 獲得圖片
        image = context.tasker.controller.cached_image

        # OCR ON
        reco_detail = context.run_recognition(
           "OCRTask", image, pipeline_override={"OCRTask": {"recognition": "OCR","roi":"辨別位置(樺石發財樹)"}} #Need to set roi
        )
        print(reco_detail) #To test
        # 檢查識別結果
        if reco_detail and reco_detail.best_result:
            ocr_text = reco_detail.best_result.text
            print(f"OCR识别结果: {ocr_text}")

            # EXCEL
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OCR结果"

            # 標題
            ws.cell(row=1, column=1, value="OCR识别结果")
            # OCR RESUALT
            ws.cell(row=2, column=1, value=ocr_text)

            # 保存Excel
            excel_file_path = "ocr_results.xlsx"
            wb.save(excel_file_path)
            print(f"OCR结果已保存到 {excel_file_path}")

        else:
            print("未获取到OCR识别结果")

        return True
    
class OCRCityAction(CustomAction):
    def run(
        self, context: Context, argv: CustomAction.RunArg
    ) -> bool:

        print(f"正在执行 OCRCityAction.run，context: {context}, argv: {argv}")

        # 獲得圖片
        image = context.tasker.controller.cached_image

        # OCR ON
        reco_detail = context.run_recognition(
           "OCRTask", image, pipeline_override={"OCRTask": {"recognition": "OCR","roi":[ 1129, 499, 62, 16]}} #Need to set roi
        )
        print(reco_detail) #To test
        # 檢查識別結果
        if reco_detail and reco_detail.best_result:
            ocr_text = reco_detail.best_result.text
            print(f"OCR识别结果: {ocr_text}")

            # EXCEL
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OCR结果"

            # 標題
            ws.cell(row=1, column=1, value=ocr_text)
            context.override_pipeline({"OCRCity":{"next":[ocr_text]}})
            # 保存Excel
            excel_file_path = "ocr_results.xlsx"
            wb.save(excel_file_path)
            print(f"OCR结果已保存到 {excel_file_path}")
    
        else:
            print("未获取到OCR识别结果")

        return True
    
if __name__ == "__main__":
    main()